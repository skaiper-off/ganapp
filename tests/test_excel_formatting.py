from openpyxl import Workbook
from utils_number import parse_input_number


def apply_formatting(ws):
    for row in range(2, ws.max_row + 1):
        peso_cell = ws.cell(row=row, column=4)
        costo_cell = ws.cell(row=row, column=5)
        if peso_cell.value is not None:
            try:
                peso_cell.value = float(peso_cell.value)
                peso_cell.number_format = '#,##0.00'
            except:
                pass
        if costo_cell.value is not None:
            try:
                costo_cell.value = float(costo_cell.value)
                costo_cell.number_format = '#,##0.00'
            except:
                pass


def test_excel_formatting_numeric_conversion():
    wb = Workbook()
    ws = wb.active
    ws.append(["Arete", "Tipo", "Fecha", "Peso", "Costo", "Notas"])

    # Insert rows with mixed formats
    ws.append(["A1", "Venta", "2025-12-20", "25,000", "25,000", "venta grande"])
    ws.append(["A2", "Pesaje", "2025-12-20", "350.5", "", ""])
    ws.append(["A3", "Otro", "2025-12-20", None, "2500.75", ""])

    # Simular conversión (como en exportar_eventos_excel)
    # Primero normalizar usando parse_input_number cuando corresponda
    for row in range(2, ws.max_row + 1):
        p = ws.cell(row=row, column=4).value
        c = ws.cell(row=row, column=5).value
        if isinstance(p, str):
            try:
                ws.cell(row=row, column=4).value = parse_input_number(p, "peso")
            except:
                pass
        if isinstance(c, str):
            try:
                ws.cell(row=row, column=5).value = parse_input_number(c, "costo")
            except:
                pass

    apply_formatting(ws)

    # Verificar que ahora sean floats (o None) y con number_format aplicado
    assert isinstance(ws.cell(row=2, column=4).value, float)
    assert isinstance(ws.cell(row=2, column=5).value, float)
    assert ws.cell(row=2, column=4).number_format == '#,##0.00'
    assert ws.cell(row=2, column=5).number_format == '#,##0.00'

    # Row 3 checks
    assert isinstance(ws.cell(row=3, column=4).value, float)
    assert ws.cell(row=3, column=5).value is None or isinstance(ws.cell(row=3, column=5).value, float)

    # Row 4 checks
    assert ws.cell(row=4, column=4).value is None
    assert isinstance(ws.cell(row=4, column=5).value, float)
